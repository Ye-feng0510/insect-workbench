"""Excel 模板路由测试。"""
import io
import json
import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.models import ExcelTemplate, SpecimenRecord, STATUS_COMPLETED
from app.services import excel_service, template_service

TEST_TEMPLATE = Path(__file__).resolve().parent.parent.parent / "test-data" / "示例模板表.xlsx"


@pytest.fixture
def client():
    """使用内存 SQLite 的测试客户端。"""
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    TestSession = sessionmaker(bind=engine)

    def override_get_db():
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    client.test_session_factory = TestSession
    yield client
    app.dependency_overrides.clear()


@pytest.fixture
def uploaded_template(client):
    """上传示例模板并返回模板信息。"""
    with open(TEST_TEMPLATE, "rb") as f:
        resp = client.post(
            "/api/templates/upload",
            files={"file": ("示例模板表.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 200
    return resp.json()


class TestUpload:
    """模板上传。"""

    def test_upload_success(self, uploaded_template):
        t = uploaded_template
        assert t["original_filename"] == "示例模板表.xlsx"
        assert t["is_active"] is True
        assert t["id"] > 0

    def test_upload_rejects_non_xlsx(self, client):
        resp = client.post(
            "/api/templates/upload",
            files={"file": ("test.csv", io.BytesIO(b"a,b,c"), "text/csv")},
        )
        assert resp.status_code == 400
        assert ".xlsx" in resp.json()["detail"]


class TestSheets:
    """工作表列表。"""

    def test_list_sheets(self, client, uploaded_template):
        tid = uploaded_template["id"]
        resp = client.get(f"/api/templates/{tid}/sheets")
        assert resp.status_code == 200
        sheets = resp.json()
        names = [s["name"] for s in sheets]
        assert "示例" in names
        assert "实际要录入的表格" in names


class TestInspect:
    """表头检测和字段映射。"""

    def test_auto_detect_header_and_mapping(self, client, uploaded_template):
        tid = uploaded_template["id"]
        resp = client.post(
            f"/api/templates/{tid}/inspect",
            params={"sheet_name": "实际要录入的表格"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["sheet_name"] == "实际要录入的表格"
        assert data["detected_header_row"] == 1
        mapping = data["field_mapping"]
        # 14 个字段应该全部自动匹配
        assert len(mapping) == 14
        assert mapping["中名"] == "E"
        assert mapping["图像"] == "AE"
        assert mapping["采集日期"] == "AJ"
        assert mapping["鉴定人"] == "AM"
        assert len(data["unmatched"]) == 0


class TestSaveMapping:
    """保存字段映射和 base_write_row 计算。"""

    def test_save_mapping_calculates_base_write_row(self, client, uploaded_template):
        tid = uploaded_template["id"]
        # 先 inspect 获取映射
        resp = client.post(
            f"/api/templates/{tid}/inspect",
            params={"sheet_name": "实际要录入的表格"},
        )
        mapping = resp.json()["field_mapping"]

        # 保存映射
        resp = client.put(
            f"/api/templates/{tid}/mapping",
            json={
                "target_sheet": "实际要录入的表格",
                "header_row": 1,
                "start_row": 2,
                "style_source_row": 2,
                "field_mapping": mapping,
            },
        )
        assert resp.status_code == 200
        data = resp.json()
        # 未映射列在第7行仍有模板数据，必须在所有客户数据之后写入。
        assert data["base_write_row"] == 8
        assert data["target_sheet"] == "实际要录入的表格"

    def test_save_mapping_appends_after_data_beyond_blank_gaps(
        self, client, uploaded_template
    ):
        tid = uploaded_template["id"]
        inspected = client.post(
            f"/api/templates/{tid}/inspect",
            params={"sheet_name": "示例"},
        ).json()

        resp = client.put(
            f"/api/templates/{tid}/mapping",
            json={
                "target_sheet": "示例",
                "header_row": inspected["detected_header_row"],
                "start_row": 2,
                "style_source_row": 2,
                "field_mapping": inspected["field_mapping"],
            },
        )

        assert resp.status_code == 200
        assert resp.json()["base_write_row"] == 11

    def test_save_mapping_requires_zhongming_and_tuxiang(self, client, uploaded_template):
        """中名和图像必须有列映射。"""
        tid = uploaded_template["id"]
        resp = client.put(
            f"/api/templates/{tid}/mapping",
            json={
                "target_sheet": "实际要录入的表格",
                "header_row": 1,
                "start_row": 2,
                "style_source_row": 2,
                "field_mapping": {"Phylum": "G"},  # 缺中名和图像
            },
        )
        assert resp.status_code == 400
        assert "中名" in resp.json()["detail"]


class TestGetMapping:
    """获取已保存配置。"""

    def test_get_current_template(self, client, uploaded_template):
        resp = client.get("/api/templates/current")
        assert resp.status_code == 200
        data = resp.json()
        assert data["id"] == uploaded_template["id"]

    def test_get_current_repairs_legacy_unsafe_write_row(
        self, client, uploaded_template, monkeypatch
    ):
        tid = uploaded_template["id"]
        inspected = client.post(
            f"/api/templates/{tid}/inspect",
            params={"sheet_name": "示例"},
        ).json()
        calculate = template_service.calculate_base_write_row
        monkeypatch.setattr(
            template_service,
            "calculate_base_write_row",
            lambda *_args, **_kwargs: 5,
        )
        saved = client.put(
            f"/api/templates/{tid}/mapping",
            json={
                "target_sheet": "示例",
                "header_row": inspected["detected_header_row"],
                "start_row": 2,
                "style_source_row": 2,
                "field_mapping": inspected["field_mapping"],
            },
        )
        assert saved.json()["base_write_row"] == 5
        monkeypatch.setattr(
            template_service,
            "calculate_base_write_row",
            calculate,
        )

        current = client.get("/api/templates/current")

        assert current.status_code == 200
        assert current.json()["base_write_row"] == 11

    def test_get_current_none(self, client):
        """没有上传模板时返回 null。"""
        resp = client.get("/api/templates/current")
        assert resp.status_code == 200
        assert resp.json() is None


class TestDirectExport:
    """直接导出也必须修复旧版不安全写入行。"""

    def test_preview_and_export_preserve_unmapped_customer_data(
        self, client, tmp_path, monkeypatch
    ):
        workbook_path = tmp_path / "blank-gap-template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "标本表"
        sheet["A1"] = "中名"
        sheet["B1"] = "图像"
        sheet["A2"] = "已有标本"
        sheet["B2"] = "OLD-001"
        sheet["C8"] = "仅存在于未映射列的甲方数据"
        sheet["A12"].fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00",
        )
        workbook.save(workbook_path)
        workbook.close()

        with client.test_session_factory() as db:
            db.add(
                ExcelTemplate(
                    owner_id=1,
                    original_filename=workbook_path.name,
                    stored_path=str(workbook_path),
                    target_sheet="标本表",
                    header_row=1,
                    start_row=2,
                    base_write_row=3,
                    style_source_row=2,
                    field_mapping_json=json.dumps(
                        {"中名": "A", "图像": "B"}, ensure_ascii=False
                    ),
                    is_active=True,
                )
            )
            db.add(
                SpecimenRecord(
                    owner_id=1,
                    zhongming="新导出标本",
                    tuxiang="NEW-009",
                    status=STATUS_COMPLETED,
                )
            )
            db.commit()

        monkeypatch.setattr(excel_service, "EXPORTS_DIR", tmp_path / "exports")

        preview_response = client.get("/api/excel/preview?mode=all")

        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["base_write_row"] == 9
        customer_row = next(
            row for row in preview["rows"] if row["excel_row"] == 8
        )
        assert customer_row["values"]["列C"] == "仅存在于未映射列的甲方数据"
        completed_row = next(
            row for row in preview["rows"] if row["status"] == "completed"
        )
        assert completed_row["excel_row"] == 9

        response = client.post("/api/export/excel")

        assert response.status_code == 200
        exported_path = (
            excel_service.EXPORTS_DIR / response.json()["filename"]
        )
        exported = load_workbook(exported_path, data_only=True)
        exported_sheet = exported["标本表"]
        assert exported_sheet["C8"].value == "仅存在于未映射列的甲方数据"
        assert exported_sheet["A9"].value == "新导出标本"
        assert exported_sheet["B9"].value == "NEW-009"
        assert exported_sheet["A12"].value is None
        assert exported_sheet["A12"].fill.fill_type == "solid"
        exported.close()

        with client.test_session_factory() as db:
            assert db.query(ExcelTemplate).one().base_write_row == 9


class TestTestMapping:
    """测试模板配置。"""

    def test_test_mapping_after_save(self, client, uploaded_template):
        tid = uploaded_template["id"]
        # inspect + save
        resp = client.post(
            f"/api/templates/{tid}/inspect",
            params={"sheet_name": "实际要录入的表格"},
        )
        mapping = resp.json()["field_mapping"]
        client.put(
            f"/api/templates/{tid}/mapping",
            json={
                "target_sheet": "实际要录入的表格",
                "header_row": 1,
                "start_row": 2,
                "style_source_row": 2,
                "field_mapping": mapping,
            },
        )
        # test
        resp = client.post(f"/api/templates/{tid}/test")
        assert resp.status_code == 200
        data = resp.json()
        assert data["sheet_name"] == "实际要录入的表格"
        assert data["base_write_row"] == 8
        assert data["mapped_count"] == 14
