"""Excel 实时预览服务。

清单第 12.4 节:
- preview_service 负责把模板已有行与数据库中的已完成记录合并成预览数据。
- 与 excel_service 共享同一个"记录转 Excel 行"的转换函数。
- 临时行(草稿)由前端叠加,后端只返回已完成记录。
"""
from __future__ import annotations

import json
from datetime import datetime
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy.orm import Session

from app.field_mapping import ALL_TARGET_FIELDS
from app.models import ExcelTemplate, SpecimenRecord, STATUS_COMPLETED
from app.services import recognition_service, template_service


def _get_active_template_or_400(db: Session, owner_id: int) -> ExcelTemplate:
    """获取活跃模板,不存在则 400。"""
    template = (
        db.query(ExcelTemplate)
        .filter(
            ExcelTemplate.owner_id == owner_id,
            ExcelTemplate.is_active == True,  # noqa: E712
        )
        .first()
    )
    if template is None or not template.target_sheet:
        raise HTTPException(
            status_code=400,
            detail="尚未配置 Excel 模板,请先在“模板与导出”页面上传模板并保存字段映射",
        )
    return template


def get_preview(
    db: Session,
    mode: str = "target",
    limit: int = 100,
    owner_id: int | None = None,
) -> dict[str, Any]:
    """生成 Excel 预览数据。

    清单第 8.6 节返回结构。
    """
    if owner_id is None:
        raise ValueError("owner_id is required")
    template = _get_active_template_or_400(db, owner_id)
    field_mapping = json.loads(template.field_mapping_json)

    # 读取模板表头和已有数据
    try:
        wb = load_workbook(
            template_service.resolve_template_path(template),
            read_only=True,
            data_only=True,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}")

    if template.target_sheet not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=400, detail=f"工作表 '{template.target_sheet}' 不存在")

    ws = wb[template.target_sheet]

    # 确定要显示的列
    if mode == "all":
        # 全部列:读取表头行的所有非空列
        columns = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(template.header_row, c).value
            letter = get_column_letter(c)
            field_name = str(v).strip() if v else f"列{letter}"
            columns.append({"letter": letter, "field": field_name})
    else:
        # 仅目标字段
        columns = [{"letter": letter, "field": field} for field, letter in field_mapping.items()]

    # 读取模板已有数据行(header_row+1 到 base_write_row-1)
    template_rows: list[dict[str, Any]] = []
    for r in range(template.header_row + 1, template.base_write_row):
        values: dict[str, str] = {}
        for col in columns:
            col_idx = column_index_from_string(col["letter"])
            v = ws.cell(r, col_idx).value
            if v is not None:
                # 日期格式化
                if isinstance(v, datetime):
                    values[col["field"]] = v.strftime("%Y-%m-%d")
                else:
                    values[col["field"]] = str(v)
            else:
                values[col["field"]] = ""
        template_rows.append({
            "excel_row": r,
            "record_id": None,
            "status": "template",
            "values": values,
        })

    wb.close()

    # 读取已完成记录(按 id 升序)
    completed_records = (
        db.query(SpecimenRecord)
        .filter(
            SpecimenRecord.owner_id == owner_id,
            SpecimenRecord.status == STATUS_COMPLETED,
        )
        .order_by(SpecimenRecord.id.asc())
        .all()
    )

    record_rows: list[dict[str, Any]] = []
    for idx, record in enumerate(completed_records):
        excel_row = template.base_write_row + idx
        fields = recognition_service.record_to_fields(record)
        # 只保留预览模式中显示的字段
        if mode == "target":
            values = {col["field"]: fields.get(col["field"], "") for col in columns}
        else:
            # 全部列模式:同样只填目标字段,其他列留空
            values = {}
            for col in columns:
                if col["field"] in fields:
                    values[col["field"]] = fields[col["field"]]
                else:
                    values[col["field"]] = ""
        record_rows.append({
            "excel_row": excel_row,
            "record_id": record.id,
            "status": "completed",
            "values": values,
        })

    all_rows = template_rows + record_rows

    # 计算统计信息
    completed_count = len(completed_records)
    latest_write_row = template.base_write_row + completed_count - 1 if completed_count > 0 else None
    next_write_row = template.base_write_row + completed_count

    return {
        "sheet_name": template.target_sheet,
        "mode": mode,
        "header_row": template.header_row,
        "base_write_row": template.base_write_row,
        "columns": columns,
        "rows": all_rows,
        "completed_count": completed_count,
        "latest_write_row": latest_write_row,
        "next_write_row": next_write_row,
        "last_updated": datetime.now().isoformat(),
    }
