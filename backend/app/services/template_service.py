"""Excel 模板处理服务。

清单第 12 节要求:
- 上传 .xlsx,保存到 data/templates/,数据库记录
- 读取所有工作表名称
- 扫描前 20 行寻找表头(能匹配多个目标字段的行)
- 自动生成字段到列的映射
- 计算并保存 base_write_row
"""
from __future__ import annotations

import json
import shutil
import uuid
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import TEMPLATES_DIR
from app.field_mapping import ALL_TARGET_FIELDS
from app.models import ExcelTemplate


class TemplateError(Exception):
    """模板处理错误。"""


def upload_template(db: Session, file: UploadFile) -> ExcelTemplate:
    """上传 Excel 模板,保存文件并创建数据库记录。

    新上传的模板自动设为活跃,旧模板取消活跃。
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 格式的 Excel 文件")

    # 生成唯一文件名防冲突
    suffix = uuid.uuid4().hex[:8]
    stored_name = f"template_{suffix}.xlsx"
    stored_path = TEMPLATES_DIR / stored_name

    # 保存文件
    try:
        with stored_path.open("wb") as f:
            shutil.copyfileobj(file.file, f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"保存模板文件失败: {e}") from e

    # 验证文件可读
    try:
        wb = load_workbook(stored_path, read_only=True)
        wb.close()
    except Exception as e:
        stored_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"Excel 文件损坏或格式不正确: {e}") from e

    # 取消旧模板活跃状态
    db.query(ExcelTemplate).filter(ExcelTemplate.is_active == True).update(  # noqa: E712
        {ExcelTemplate.is_active: False}
    )

    # 创建新记录
    template = ExcelTemplate(
        original_filename=file.filename,
        stored_path=str(stored_path),
        is_active=True,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


def get_active_template(db: Session) -> ExcelTemplate | None:
    """获取当前活跃模板。"""
    return (
        db.query(ExcelTemplate)
        .filter(ExcelTemplate.is_active == True)  # noqa: E712
        .first()
    )


def get_template_or_404(db: Session, template_id: int) -> ExcelTemplate:
    """按 ID 获取模板,不存在则 404。"""
    obj = db.get(ExcelTemplate, template_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="模板不存在")
    return obj


def list_sheets(template: ExcelTemplate) -> list[dict[str, Any]]:
    """读取模板所有工作表名称及尺寸。"""
    try:
        wb = load_workbook(template.stored_path, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}") from e

    sheets = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        sheets.append({"name": sn, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return sheets


def inspect_template(
    template: ExcelTemplate,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> dict[str, Any]:
    """检查模板工作表,自动检测表头行和字段映射。

    清单 12.2:扫描前 20 行寻找表头,能匹配多个目标字段的行作为候选。
    """
    try:
        wb = load_workbook(template.stored_path, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}") from e

    target_sheet = sheet_name or template.target_sheet
    if not target_sheet or target_sheet not in wb.sheetnames:
        # 默认取第一个工作表
        target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]

    max_scan_row = min(20, ws.max_row)

    # 自动检测表头行: 匹配最多目标字段的行
    if header_row is None:
        best_row = 1
        best_matches = 0
        for r in range(1, max_scan_row + 1):
            matches = _count_target_matches(ws, r)
            if matches > best_matches:
                best_matches = matches
                best_row = r
        detected_header_row = best_row
    else:
        detected_header_row = header_row

    # 构建字段映射
    field_mapping = _build_field_mapping(ws, detected_header_row)

    wb.close()

    return {
        "sheet_name": target_sheet,
        "detected_header_row": detected_header_row,
        "field_mapping": field_mapping,
        "unmatched": [
            f for f in ALL_TARGET_FIELDS if f not in field_mapping
        ],
    }


def _count_target_matches(ws, row: int) -> int:
    """计算某行匹配了多少个目标字段。"""
    header_values = set()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v is not None:
            header_values.add(str(v).strip())
    return sum(1 for f in ALL_TARGET_FIELDS if f in header_values)


def _build_field_mapping(ws, header_row: int) -> dict[str, str]:
    """从表头行构建字段->列字母映射。"""
    mapping: dict[str, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            field_name = str(v).strip()
            if field_name in ALL_TARGET_FIELDS:
                mapping[field_name] = get_column_letter(c)
    return mapping


def calculate_base_write_row(
    template: ExcelTemplate,
    sheet_name: str,
    start_row: int,
    field_mapping: dict[str, str],
) -> int:
    """计算 base_write_row(清单 12.2 节)。

    从 start_row 开始,在"图像"列中查找第一个可写空白行;
    若没有空白行,则使用当前最后一个已用行的下一行。
    """
    tuxiang_letter = field_mapping.get("图像")
    if not tuxiang_letter:
        # 没有图像列映射,直接用 start_row
        return start_row

    try:
        wb = load_workbook(template.stored_path, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}") from e

    ws = wb[sheet_name]
    # 将列字母转为列序号
    from openpyxl.utils import column_index_from_string

    img_col = column_index_from_string(tuxiang_letter)

    first_blank = None
    last_used = start_row - 1

    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, img_col).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            if first_blank is None:
                first_blank = r
        else:
            last_used = r

    wb.close()

    if first_blank is not None:
        return first_blank
    return last_used + 1


def save_mapping(
    db: Session,
    template: ExcelTemplate,
    target_sheet: str,
    header_row: int,
    start_row: int,
    style_source_row: int,
    field_mapping: dict[str, str],
) -> ExcelTemplate:
    """保存字段映射配置,含 base_write_row 计算。"""
    # 校验目标工作表存在
    try:
        wb = load_workbook(template.stored_path, read_only=True)
        if target_sheet not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"工作表 '{target_sheet}' 不存在,可用: {wb.sheetnames}",
            )
        wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}") from e

    # 校验必填字段映射(中名和图像必须有)
    missing_required = []
    if "中名" not in field_mapping:
        missing_required.append("中名")
    if "图像" not in field_mapping:
        missing_required.append("图像")
    if missing_required:
        raise HTTPException(
            status_code=400,
            detail=f"以下必填字段缺少列映射: {'、'.join(missing_required)}",
        )

    # 计算 base_write_row
    base_write_row = calculate_base_write_row(
        template, target_sheet, start_row, field_mapping
    )

    template.target_sheet = target_sheet
    template.header_row = header_row
    template.start_row = start_row
    template.style_source_row = style_source_row
    template.field_mapping_json = json.dumps(field_mapping, ensure_ascii=False)
    template.base_write_row = base_write_row
    db.commit()
    db.refresh(template)
    return template


def test_mapping(template: ExcelTemplate) -> dict[str, Any]:
    """测试模板配置:验证映射能正确读取表头和模板数据。"""
    if not template.target_sheet:
        raise HTTPException(status_code=400, detail="尚未配置目标工作表")

    field_mapping = json.loads(template.field_mapping_json) if template.field_mapping_json else {}

    try:
        wb = load_workbook(template.stored_path, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取模板失败: {e}") from e

    if template.target_sheet not in wb.sheetnames:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail=f"工作表 '{template.target_sheet}' 不存在",
        )

    ws = wb[template.target_sheet]

    # 读取表头行验证
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(template.header_row, c).value
        if v is not None:
            headers[get_column_letter(c)] = str(v).strip()

    # 读取 base_write_row 前的模板数据示例(前 3 行)
    sample_rows = []
    from openpyxl.utils import column_index_from_string

    for r in range(
        max(template.header_row + 1, template.start_row),
        min(template.base_write_row, template.header_row + 4),
    ):
        row_data = {}
        for field, letter in field_mapping.items():
            col = column_index_from_string(letter)
            v = ws.cell(r, col).value
            row_data[field] = "" if v is None else str(v)
        sample_rows.append({"excel_row": r, "values": row_data})

    wb.close()

    return {
        "sheet_name": template.target_sheet,
        "header_row": template.header_row,
        "base_write_row": template.base_write_row,
        "style_source_row": template.style_source_row,
        "field_mapping": field_mapping,
        "mapped_count": len(field_mapping),
        "unmapped": [f for f in ALL_TARGET_FIELDS if f not in field_mapping],
        "sample_rows": sample_rows,
    }


def template_to_info(t: ExcelTemplate) -> dict[str, Any]:
    """将 ORM 对象转为 API 响应字典。"""
    mapping = json.loads(t.field_mapping_json) if t.field_mapping_json else {}
    return {
        "id": t.id,
        "original_filename": t.original_filename,
        "target_sheet": t.target_sheet,
        "header_row": t.header_row,
        "start_row": t.start_row,
        "base_write_row": t.base_write_row,
        "style_source_row": t.style_source_row,
        "field_mapping": mapping,
        "is_active": t.is_active,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }
