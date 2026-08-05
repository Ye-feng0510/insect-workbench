"""Excel 导出服务。

清单第 12.3 节导出规则:
1. 使用原模板创建副本
2. 打开目标工作表
3. 读取 base_write_row
4. 按 id 升序读取已完成记录
5. 第 index 条记录(从0开始)的行号 = base_write_row + index
6. 只写入已映射的目标字段
7. 保留 base_write_row 之前的模板原有行、其他列和格式
8. 从格式来源行复制样式
9. 采集日期写为真正的 Excel 日期,yyyy-mm-dd 格式
10. 图像编号以文本格式写入
11. 导出到 data/exports/
12. 返回下载地址
13. 不覆盖原模板
"""
from __future__ import annotations

import json
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy.orm import Session

from app.config import EXPORTS_DIR
from app.field_mapping import FIELD_TO_COLUMN
from app.models import (
    ExcelTemplate,
    ExportArtifact,
    SpecimenRecord,
    STATUS_COMPLETED,
)
from app.services import recognition_service, template_service


def _get_active_template_or_400(db: Session, owner_id: int) -> ExcelTemplate:
    """获取活跃模板。"""
    template = template_service.get_active_template(db, owner_id)
    if template is None or not template.target_sheet:
        raise HTTPException(
            status_code=400,
            detail="尚未配置 Excel 模板,请先上传模板并保存字段映射",
        )
    return template


def get_export_summary(db: Session, owner_id: int) -> dict[str, Any]:
    """获取导出汇总信息。"""
    template = _get_active_template_or_400(db, owner_id)

    completed_count = (
        db.query(SpecimenRecord)
        .filter(
            SpecimenRecord.owner_id == owner_id,
            SpecimenRecord.status == STATUS_COMPLETED,
        )
        .count()
    )
    awaiting_count = (
        db.query(SpecimenRecord)
        .filter(
            SpecimenRecord.owner_id == owner_id,
            SpecimenRecord.status == "awaiting_confirmation",
        )
        .count()
    )

    return {
        "completed_count": completed_count,
        "awaiting_confirmation_count": awaiting_count,
        "template_name": template.original_filename,
        "target_sheet": template.target_sheet,
        "start_write_row": template.base_write_row,
    }


def export_excel(
    db: Session, owner_id: int, actor_user_id: int
) -> dict[str, Any]:
    """生成导出 Excel 文件。

    清单第 12.3 节完整流程。
    """
    template = _get_active_template_or_400(db, owner_id)
    field_mapping = json.loads(template.field_mapping_json)

    # 读取已完成记录(按 id 升序)
    completed_records = (
        db.query(SpecimenRecord)
        .filter(
            SpecimenRecord.owner_id == owner_id,
            SpecimenRecord.status == STATUS_COMPLETED,
        )
        .order_by(SpecimenRecord.id.asc())
        .all()
    )

    if not completed_records:
        raise HTTPException(status_code=400, detail="没有已完成的记录,无法导出")

    # 复制模板到导出目录
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    export_filename = (
        f"昆虫标本信息_{owner_id}_{now.strftime('%Y%m%d_%H%M%S')}_"
        f"{uuid.uuid4().hex[:8]}.xlsx"
    )
    export_path = EXPORTS_DIR / export_filename
    shutil.copy2(template_service.resolve_template_path(template), export_path)

    # 打开副本进行写入
    try:
        wb = load_workbook(export_path)
    except Exception as e:
        export_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"打开模板副本失败: {e}")

    if template.target_sheet not in wb.sheetnames:
        wb.close()
        export_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=400,
            detail=f"工作表 '{template.target_sheet}' 不存在",
        )

    ws = wb[template.target_sheet]

    # 读取格式来源行的样式(用于复制到新写入行)
    style_row = template.style_source_row
    style_cells: dict[str, Any] = {}
    for field, letter in field_mapping.items():
        col_idx = column_index_from_string(letter)
        src_cell = ws.cell(style_row, col_idx)
        style_cells[field] = {
            "font": src_cell.font.copy() if src_cell.font else Font(),
            "fill": src_cell.fill.copy() if src_cell.fill else PatternFill(),
            "border": src_cell.border.copy() if src_cell.border else Border(),
            "alignment": src_cell.alignment.copy() if src_cell.alignment else Alignment(),
            "number_format": src_cell.number_format,
        }

    # 写入记录
    for idx, record in enumerate(completed_records):
        excel_row = template.base_write_row + idx
        fields = recognition_service.record_to_fields(record)

        for field, letter in field_mapping.items():
            col_idx = column_index_from_string(letter)
            cell = ws.cell(excel_row, col_idx)
            value = fields.get(field, "")

            # 应用格式来源行的样式
            if field in style_cells:
                s = style_cells[field]
                cell.font = s["font"]
                cell.fill = s["fill"]
                cell.border = s["border"]
                cell.alignment = s["alignment"]
                cell.number_format = s["number_format"]

            # 特殊格式处理
            if field == "采集日期" and value:
                # 采集日期写为真正的 Excel 日期
                try:
                    dt = datetime.strptime(value, "%Y-%m-%d")
                    cell.value = dt.date()
                    cell.number_format = "yyyy-mm-dd"
                except ValueError:
                    cell.value = value
            elif field == "图像" and value:
                # 图像编号以文本格式写入
                cell.value = str(value)
                cell.number_format = "@"
            else:
                cell.value = str(value) if value else None

    # 保存
    try:
        wb.save(export_path)
    except PermissionError:
        wb.close()
        export_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=500,
            detail="导出文件被其他程序占用,请关闭 Excel 后重试",
        )
    finally:
        wb.close()

    db.add(
        ExportArtifact(
            owner_id=owner_id,
            filename=export_filename,
            stored_path=str(export_path),
            created_by_user_id=actor_user_id,
        )
    )
    db.commit()
    return {
        "filename": export_filename,
        "download_url": f"/api/export/download/{export_filename}",
        "record_count": len(completed_records),
    }


def get_export_file_path(db: Session, filename: str, owner_id: int) -> Path:
    """获取导出文件的安全路径。"""
    safe_name = Path(filename).name
    artifact = db.query(ExportArtifact).filter(
        ExportArtifact.owner_id == owner_id,
        ExportArtifact.filename == safe_name,
    ).first()
    if artifact is None:
        raise HTTPException(status_code=404, detail="导出文件不存在")
    export_path = Path(artifact.stored_path)
    if not export_path.is_file():
        export_path = EXPORTS_DIR / safe_name
    if not export_path.exists():
        raise HTTPException(status_code=404, detail="导出文件不存在")
    return export_path
