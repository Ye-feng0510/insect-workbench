"""Small, explicit SQLite schema migration and bootstrap system."""
from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from sqlalchemy import inspect, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from app.auth import hash_password
from app.config import EXPORTS_DIR, settings
from app.database import Base
from app.models import ExcelTemplate, ExportArtifact, ROLE_ADMIN, User

LATEST_SCHEMA_VERSION = 6


def _validate_new_admin_credentials(username: str, password: str) -> None:
    if not username or not password:
        raise RuntimeError(
            "首次启动必须设置 INSECT_BOOTSTRAP_ADMIN_USERNAME 和 "
            "INSECT_BOOTSTRAP_ADMIN_PASSWORD（密码至少 12 位）"
        )
    if len(password) < 12:
        raise RuntimeError("密码至少需要 12 个字符")


def bootstrap_admin(db: Session, *, create: bool = True) -> User | None:
    username = settings.bootstrap_admin_username.strip()
    password = settings.bootstrap_admin_password
    if username:
        configured = db.query(User).filter(User.username == username).first()
        if configured is not None:
            if configured.role != ROLE_ADMIN or not configured.is_active:
                raise RuntimeError("配置的启动管理员账号不是已启用的管理员")
            return configured
    existing_admin = (
        db.query(User)
        .filter(User.role == ROLE_ADMIN, User.is_active.is_(True))
        .order_by(User.id.asc())
        .first()
    )
    if existing_admin is not None:
        return existing_admin
    _validate_new_admin_credentials(username, password)
    if not create:
        return None
    try:
        password_hash = hash_password(password)
    except ValueError as exc:
        raise RuntimeError(str(exc)) from exc
    admin = User(
        username=username,
        password_hash=password_hash,
        role=ROLE_ADMIN,
        workflow_quota=None,
        is_active=True,
    )
    db.add(admin)
    db.commit()
    db.refresh(admin)
    return admin


def _preflight_bootstrap_admin(
    engine: Engine, existing_tables: set[str]
) -> None:
    if "users" not in existing_tables:
        _validate_new_admin_credentials(
            settings.bootstrap_admin_username.strip(),
            settings.bootstrap_admin_password,
        )
        return
    with Session(engine) as db:
        bootstrap_admin(db, create=False)


def _backup_database(engine: Engine) -> None:
    if engine.url.get_backend_name() != "sqlite":
        return
    path = Path(engine.url.database or "")
    if not path.exists() or path.stat().st_size == 0:
        return
    backup = path.with_name(
        f"{path.name}.backup-{datetime.now().strftime('%Y%m%d%H%M%S%f')}-"
        f"{uuid4().hex[:8]}"
    )
    shutil.copy2(path, backup)


def _columns(conn, table: str) -> set[str]:
    return {
        row[1]
        for row in conn.execute(text(f"PRAGMA table_info('{table}')")).fetchall()
    }


def _backfill_jiandingren_mappings(engine: Engine) -> None:
    """为已有模板安全补全鉴定人列映射。"""
    from openpyxl import load_workbook

    from app.services.template_service import resolve_template_path

    columns = set()
    with engine.connect() as conn:
        if "excel_templates" in inspect(engine).get_table_names():
            columns = _columns(conn, "excel_templates")
    required = {
        "stored_path",
        "target_sheet",
        "header_row",
        "field_mapping_json",
    }
    if not required.issubset(columns):
        return

    with Session(engine) as db:
        templates = db.query(ExcelTemplate).all()
        changed = False
        for template in templates:
            if not template.target_sheet or not template.header_row:
                continue
            try:
                mapping = json.loads(template.field_mapping_json or "{}")
                if "鉴定人" in mapping:
                    continue
                wb = load_workbook(resolve_template_path(template), read_only=True)
                try:
                    if template.target_sheet not in wb.sheetnames:
                        continue
                    ws = wb[template.target_sheet]
                    matches = [
                        cell.column_letter
                        for cell in ws[template.header_row]
                        if str(cell.value or "").strip() == "鉴定人"
                    ]
                finally:
                    wb.close()
                if len(matches) == 1:
                    mapping["鉴定人"] = matches[0]
                    template.field_mapping_json = json.dumps(
                        mapping, ensure_ascii=False
                    )
                    changed = True
            except Exception:
                continue
        if changed:
            db.commit()


def migrate(engine: Engine) -> None:
    """Run all migrations in order; every error aborts startup."""
    from app import models  # noqa: F401

    existing_tables = set(inspect(engine).get_table_names())
    current_version = 0
    if "schema_version" in existing_tables:
        with engine.connect() as conn:
            current_version = conn.execute(
                text("SELECT COALESCE(MAX(version), 0) FROM schema_version")
            ).scalar_one()
    needs_legacy_migration = bool(
        existing_tables
        & {"excel_templates", "specimen_records", "material_batches"}
    ) and current_version < LATEST_SCHEMA_VERSION
    _preflight_bootstrap_admin(engine, existing_tables)
    if needs_legacy_migration:
        _backup_database(engine)

    # Auth/accounting tables must exist before legacy rows can receive owners.
    for table_name in (
        "users",
        "auth_sessions",
        "quota_adjustments",
    ):
        Base.metadata.tables[table_name].create(engine, checkfirst=True)

    with Session(engine) as db:
        admin = bootstrap_admin(db)
        assert admin is not None
        admin_id = admin.id

    # Create missing tables only. Existing-table ownership changes remain explicit below.
    Base.metadata.create_all(bind=engine)
    existing_tables = set(inspect(engine).get_table_names())

    with engine.begin() as conn:
        conn.execute(
            text(
                "CREATE TABLE IF NOT EXISTS schema_version "
                "(version INTEGER PRIMARY KEY, applied_at DATETIME NOT NULL "
                "DEFAULT CURRENT_TIMESTAMP)"
            )
        )
        applied = {
            row[0]
            for row in conn.execute(text("SELECT version FROM schema_version"))
        }

        if 1 not in applied:
            for table in ("excel_templates", "specimen_records", "material_batches"):
                if table in existing_tables and "owner_id" not in _columns(conn, table):
                    conn.execute(
                        text(
                            f"ALTER TABLE {table} ADD COLUMN owner_id INTEGER "
                            "REFERENCES users(id)"
                        )
                    )
                if table in existing_tables:
                    conn.execute(
                        text(
                            f"UPDATE {table} SET owner_id=:owner "
                            "WHERE owner_id IS NULL"
                        ),
                        {"owner": admin_id},
                    )
            conn.execute(text("DROP INDEX IF EXISTS uq_specimen_tuxiang_completed"))
            conn.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS "
                    "uq_specimen_owner_tuxiang_completed "
                    "ON specimen_records (owner_id, tuxiang) "
                    "WHERE status = 'completed'"
                )
            )
            conn.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS "
                    "uq_template_owner_active ON excel_templates(owner_id) "
                    "WHERE is_active = 1"
                )
            )
            conn.execute(
                text(
                    "CREATE UNIQUE INDEX IF NOT EXISTS "
                    "uq_batch_owner_active ON material_batches(owner_id) "
                    "WHERE is_active = 1"
                )
            )
            conn.execute(
                text("INSERT INTO schema_version(version) VALUES (1)")
            )

        if 2 not in applied:
            if (
                "material_prefetch_results" in existing_tables
                and "attempt_count"
                not in _columns(conn, "material_prefetch_results")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE material_prefetch_results "
                        "ADD COLUMN attempt_count INTEGER DEFAULT 0"
                    )
                )
            if (
                "material_prefetch_results" in existing_tables
                and "next_retry_at"
                not in _columns(conn, "material_prefetch_results")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE material_prefetch_results "
                        "ADD COLUMN next_retry_at DATETIME"
                    )
                )
            conn.execute(text("INSERT INTO schema_version(version) VALUES (2)"))

        if 3 not in applied:
            if (
                "taxonomy_cache" in existing_tables
                and "owner_id" not in _columns(conn, "taxonomy_cache")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE taxonomy_cache "
                        "RENAME TO taxonomy_cache_legacy_v3"
                    )
                )
                conn.execute(
                    text(
                        "CREATE TABLE taxonomy_cache ("
                        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
                        "owner_id INTEGER NOT NULL REFERENCES users(id) "
                        "ON DELETE CASCADE,"
                        "zhongming VARCHAR(200) NOT NULL,"
                        "phylum VARCHAR(200) NOT NULL DEFAULT '',"
                        "gang VARCHAR(200) NOT NULL DEFAULT '',"
                        "klass VARCHAR(200) NOT NULL DEFAULT '',"
                        "order_field VARCHAR(200) NOT NULL DEFAULT '',"
                        "zhongwen_ke VARCHAR(200) NOT NULL DEFAULT '',"
                        "ke VARCHAR(200) NOT NULL DEFAULT '',"
                        "shu VARCHAR(200) NOT NULL DEFAULT '',"
                        "zhong VARCHAR(200) NOT NULL DEFAULT '',"
                        "created_at DATETIME DEFAULT CURRENT_TIMESTAMP,"
                        "updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,"
                        "UNIQUE(owner_id, zhongming)"
                        ")"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO taxonomy_cache ("
                        "id,owner_id,zhongming,phylum,gang,klass,order_field,"
                        "zhongwen_ke,ke,shu,zhong,created_at,updated_at"
                        ") SELECT id,:owner,zhongming,phylum,gang,klass,"
                        "order_field,zhongwen_ke,ke,shu,zhong,created_at,"
                        "updated_at FROM taxonomy_cache_legacy_v3"
                    ),
                    {"owner": admin_id},
                )
                conn.execute(text("DROP TABLE taxonomy_cache_legacy_v3"))
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_taxonomy_cache_owner_id "
                    "ON taxonomy_cache(owner_id)"
                )
            )
            conn.execute(
                text(
                    "CREATE INDEX IF NOT EXISTS ix_taxonomy_cache_zhongming "
                    "ON taxonomy_cache(zhongming)"
                )
            )
            conn.execute(text("INSERT INTO schema_version(version) VALUES (3)"))

        if 4 not in applied:
            if (
                "specimen_records" in existing_tables
                and "jiandingren" not in _columns(conn, "specimen_records")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE specimen_records ADD COLUMN "
                        "jiandingren VARCHAR(200) NOT NULL DEFAULT ''"
                    )
                )
            if (
                "specimen_records" in existing_tables
                and "ocr_result_json" not in _columns(conn, "specimen_records")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE specimen_records ADD COLUMN "
                        "ocr_result_json TEXT NOT NULL DEFAULT ''"
                    )
                )
            conn.execute(text("INSERT INTO schema_version(version) VALUES (4)"))

        if 5 not in applied:
            additions = {
                "scientific_name": "VARCHAR(300) NOT NULL DEFAULT ''",
                "scientific_name_authorship": "VARCHAR(300) NOT NULL DEFAULT ''",
                "subfamily": "VARCHAR(200) NOT NULL DEFAULT ''",
                "tribe": "VARCHAR(200) NOT NULL DEFAULT ''",
                "subgenus": "VARCHAR(200) NOT NULL DEFAULT ''",
                "taxonomy_verification_json": "TEXT NOT NULL DEFAULT ''",
            }
            if "specimen_records" in existing_tables:
                record_columns = _columns(conn, "specimen_records")
                for column, definition in additions.items():
                    if column not in record_columns:
                        conn.execute(
                            text(
                                f"ALTER TABLE specimen_records ADD COLUMN "
                                f"{column} {definition}"
                            )
                        )
            conn.execute(text("INSERT INTO schema_version(version) VALUES (5)"))

        if 6 not in applied:
            if (
                "workflow_sessions" in existing_tables
                and "result_record_id"
                not in _columns(conn, "workflow_sessions")
            ):
                conn.execute(
                    text(
                        "ALTER TABLE workflow_sessions ADD COLUMN "
                        "result_record_id INTEGER REFERENCES "
                        "specimen_records(id) ON DELETE SET NULL"
                    )
                )
            conn.execute(text("INSERT INTO schema_version(version) VALUES (6)"))

    # Creates only missing tables; ownership changes above never rely on create_all.
    Base.metadata.create_all(bind=engine)
    _backfill_jiandingren_mappings(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                "CREATE INDEX IF NOT EXISTS "
                "ix_workflow_sessions_result_record_id "
                "ON workflow_sessions(result_record_id)"
            )
        )
        conn.execute(
            text(
                "CREATE INDEX IF NOT EXISTS ix_material_queue "
                "ON material_items (batch_id, status, sequence)"
            )
        )

    # Preserve and authorize pre-feature export files for the bootstrap admin.
    with Session(engine) as db:
        known = {
            row[0]
            for row in db.query(ExportArtifact.filename).all()
        }
        for path in EXPORTS_DIR.glob("*.xlsx"):
            if path.name not in known:
                db.add(
                    ExportArtifact(
                        owner_id=admin_id,
                        filename=path.name,
                        stored_path=str(path),
                        created_by_user_id=admin_id,
                    )
                )
        db.commit()
